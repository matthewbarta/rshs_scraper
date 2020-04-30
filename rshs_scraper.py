import requests, bs4, re, openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = '99s by Account Type'

#Skill Array
skills = ['Attack', 'Defence', 'Strength', 'Hitpoints', 'Ranged', 'Prayer', 'Magic', 'Cooking', 'Woodcutting', 'Fletching', 'Fishing', 'Firemaking', 'Crafting', 'Smithing', 'Mining', 'Herblore', 'Agility', 'Thieving', 'Slayer', 'Farming', 'Runecraft', 'Hunter', 'Construction']
modes = []
#Change this number as time goes on to reflect what highest possible 99 could be.
highestNumber = [300000, 150000, 30000, 25000]
lowestNumber = [20000, 25, 25, 25]

# Gamemode specific url modifiers
gametype_url = ['', '_ironman','_ultimate','_hardcore_ironman']

# For printing
gamemode = ['Regular', 'Ironman', 'Ultimate Ironman', 'Hardcore Ironman']

# Set up the spreadsheet.
headerFont = Font(name='Times New Roman', bold=True)
# Format height and width.
sheet.column_dimensions['B'].width = 15
sheet.column_dimensions['C'].width = 20
sheet.column_dimensions['D'].width = 20
sheet.column_dimensions['E'].width = 20
sheet.column_dimensions['F'].width = 20
# Set up axis'.
sheet.cell(row=3, column=2).value = 'Skill'
sheet.cell(row=3, column=2).font = headerFont
for num, skillName in enumerate(skills):
    sheet.cell(row=num+4, column=2).value = skills[num]
    sheet.cell(row=num+4, column=2).font = headerFont
for num, mode in enumerate(gamemode):
    sheet.cell(row=2, column=num+3).value = gamemode[num]
    sheet.cell(row=2, column=num+3).font = headerFont

# Loop over each gamemode.
for gm in range(len(gametype_url)):
    # Loop over individual skills.
    rank = []
    print(gamemode[gm] + ": ")
    for skill in range(len(skills)):
        # Variables for binary search, pages to search.
        lowPage = lowestNumber[gm] / 25
        highPage = highestNumber[gm] / 25
        # Conditions for determining rank, if we hit the last person
        lastRank = 1
        lastLevel = '0'
        done = False
        # This loops over pages of an individual skill.
        while not done and lowPage <= highPage:
            currentPage = int(lowPage + (highPage - lowPage) // 2)
            #! DEBUG print("Current page: " + str(currentPage) + ", High: " + str(highPage) + ", Low: " + str(lowPage))
            # Base url
            regularAccountsUrl = f'https://secure.runescape.com/m=hiscore_oldschool{gametype_url[gm]}/overall?table={skill+1}&page={currentPage}'
            res = requests.get(regularAccountsUrl)
            res.raise_for_status()
            page = bs4.BeautifulSoup(res.text, 'html.parser')
            elems = page.select('.personal-hiscores__row')
            #Searches each individual hs page.
            for i in range(len(elems)):
                elems[i] = [x for x in elems[i].getText().splitlines() if x != '']
                currentLevel = elems[i][2]
                #Remove the commas in the rank
                currentRank = elems[i][0].replace(',', '')
                #Condition for last person w/ 99
                if currentLevel == '98' and lastLevel == '99' and lastRank == int(currentRank) - 1:
                    done = True
                    break
                #Update Rank
                else:
                    lastRank = int(currentRank)
                    lastLevel = currentLevel
            # Update page based on binary search-esque algorithm
            if lastLevel == '99':
                lowPage = currentPage + 1
            else:
                highPage = currentPage - 1
        # Add the data to an excel spreadsheet, comma-separated.
        sheet.cell(row = 4 + skill, column = 3 + gm).value = int(lastRank)
        sheet.cell(row = 4 + skill, column = 3 + gm).number_format = '#,##0'
        print(skills[skill] + ": " + str(lastRank))
wb.save('rshs_99s.xlsx')